import openpyxl

# class to read data from excel
class readexcel:

    # we defined the method as static so that we an access this method without creating any object
    @staticmethod
    # for this class and directly via the class name
    def getdatafromexcel(scenarioname):
        book =openpyxl.load_workbook("C:\\DEV\\PYTHON\\PycharmProjects\\Ecommerceframework\\TestData\\Userdata.xlsx")
        sheet = book.active

        #reading each column header and its values and storing it as key value pair in a dictonary
        dict={}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row =i, column =1).value == scenarioname:
                for j in range(2, sheet.max_column+1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return dict



